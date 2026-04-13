"""
01_download_ohlcv.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
KOSPI 200 종목 OHLCV 다운로드 (yfinance)
결과: ohlcv_cache/{ticker}.pkl  (pandas DataFrame)

설정
────
PERIOD  : 다운로드 기간 (기본 3년 — 슬라이딩 윈도우 여유분 포함)
WORKERS : 병렬 다운로드 스레드 수
"""

import os
import sys
import time
import pickle
import logging
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import pandas as pd
import yfinance as yf

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
XLSX_PATH  = BASE_DIR / "data" / "kospi200.xlsx"
CACHE_DIR  = BASE_DIR / "ohlcv_cache"
CACHE_DIR.mkdir(exist_ok=True)

# ── 다운로드 설정 ──────────────────────────────────────────────────────────────
PERIOD   = "3y"      # 3년치 (6개월 X + 슬라이딩 + 여유)
WORKERS  = 8         # 병렬 다운로드 수 (너무 높이면 yfinance rate limit)
INTERVAL = "1d"      # 일봉

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def load_tickers() -> list[dict]:
    """kospi200.xlsx → [{code, name}, ...] 200개"""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    tickers = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if code:
            tickers.append({"code": str(code).zfill(6), "name": name})
    log.info(f"KOSPI 200 로드 완료: {len(tickers)}종목")
    return tickers


def download_one(ticker_info: dict) -> tuple[str, bool, str]:
    """단일 종목 다운로드 → pkl 저장. (code, success, message) 반환"""
    code  = ticker_info["code"]
    name  = ticker_info["name"]
    yf_symbol = f"{code}.KS"
    out_path  = CACHE_DIR / f"{code}.pkl"

    # 이미 존재하면 스킵
    if out_path.exists():
        return code, True, f"{name} — 캐시 존재, 스킵"

    try:
        raw = yf.download(yf_symbol, period=PERIOD, interval=INTERVAL,
                          auto_adjust=True, progress=False)
        if isinstance(raw.columns, pd.MultiIndex):
            raw.columns = raw.columns.get_level_values(0)
        raw.columns = raw.columns.str.lower()
        # 중복 컬럼 제거 (yfinance 복수 종목 반환 대응)
        raw = raw.loc[:, ~raw.columns.duplicated(keep="first")]

        if len(raw) < 20:
            return code, False, f"{name} — 데이터 부족 ({len(raw)}행)"

        with open(out_path, "wb") as f:
            pickle.dump(raw, f)

        return code, True, f"{name} — {len(raw)}일 저장 완료"

    except Exception as e:
        return code, False, f"{name} — 오류: {e}"


def main():
    tickers = load_tickers()
    total   = len(tickers)
    success = 0
    fail    = 0

    log.info(f"다운로드 시작: {total}종목 / {WORKERS}스레드 병렬")

    with ThreadPoolExecutor(max_workers=WORKERS) as executor:
        futures = {executor.submit(download_one, t): t for t in tickers}
        for i, future in enumerate(as_completed(futures), 1):
            code, ok, msg = future.result()
            if ok:
                success += 1
                log.info(f"[{i:3}/{total}] ✓ {msg}")
            else:
                fail += 1
                log.warning(f"[{i:3}/{total}] ✗ {msg}")

            # rate limit 완화
            time.sleep(0.05)

    log.info(f"\n완료: 성공 {success} / 실패 {fail} / 전체 {total}")
    log.info(f"캐시 위치: {CACHE_DIR}")


if __name__ == "__main__":
    main()
