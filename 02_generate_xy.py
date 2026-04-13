"""
02_generate_xy.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
슬라이딩 윈도우로 X/Y 학습 페어 생성 → JSONL 저장

윈도우 구조
────────────
  ◀── X: 6개월 (약 126 거래일) ──▶◀─ Y: 10 거래일 ─▶
  [                               ][                ]
  shift = 1개월 (약 21 거래일)씩 이동

X (입력)
  - text_chart.py 로 렌더링한 텍스트 차트 (캔들 + MA20/60 + 볼린저)
  - generate_metadata() 자동 메타데이터

Y (레이블)
  - 10거래일간 등락률 (종가 기준)
  - 최대 상승폭 / 최대 낙폭 (MDD)
  - 방향 레이블: UP / DOWN / FLAT (±2% 기준)
  - 일별 OHLCV 요약 (LLM이 결과를 읽는 용도)

출력 형식 (xy_pairs/pairs.jsonl)
─────────────────────────────────
  {
    "id": "005930_2024-01-02",
    "ticker": "005930",
    "name": "삼성전자",
    "x_start": "2023-07-03",
    "x_end":   "2024-01-02",
    "y_start": "2024-01-03",
    "y_end":   "2024-01-16",
    "x_chart": "...(텍스트 차트)...",
    "x_meta":  "...(메타데이터)...",
    "y_return_pct": 3.45,
    "y_max_up":     5.12,
    "y_max_down":  -1.23,
    "y_direction":  "UP",
    "y_daily":      [{...}, ...]
  }
"""

import sys
import json
import pickle
import logging
from pathlib import Path
from datetime import timedelta

import numpy as np
import pandas as pd

# ── 경로 ──────────────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent
CACHE_DIR = BASE_DIR / "ohlcv_cache"
OUT_DIR   = BASE_DIR / "xy_pairs"
OUT_DIR.mkdir(exist_ok=True)

# text_chart.py 는 같은 폴더에 포함됨
sys.path.insert(0, str(BASE_DIR))
from text_chart import plot_text_chart, add_ma, add_bollinger, generate_metadata

# ── 윈도우 설정 ───────────────────────────────────────────────────────────────
X_DAYS    = 126   # 6개월 ≈ 126 거래일
Y_DAYS    = 10    # 미래 10 거래일
SHIFT     = 21    # 1개월 ≈ 21 거래일씩 shift
FLAT_THR  = 2.0   # ±2% 미만이면 FLAT

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ── 유틸 ──────────────────────────────────────────────────────────────────────

def load_ticker_meta() -> dict:
    """kospi200.xlsx → {code: name} 딕셔너리"""
    import openpyxl
    xlsx = BASE_DIR / "data" / "kospi200.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active
    meta = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if code:
            meta[str(code).zfill(6)] = name
    return meta


def compute_y_stats(y_df: pd.DataFrame) -> dict:
    """10거래일 Y 구간 통계 산출"""
    close0 = float(y_df["close"].iloc[0])   # Y 시작 종가 (기준)
    closes = y_df["close"].astype(float)

    # 기준 대비 등락률 시리즈
    rets = (closes / close0 - 1.0) * 100.0

    y_return_pct = float(rets.iloc[-1])       # 10일 후 최종 등락률
    y_max_up     = float(rets.max())           # 구간 내 최대 상승
    y_max_down   = float(rets.min())           # 구간 내 최대 낙폭

    if y_return_pct >= FLAT_THR:
        direction = "UP"
    elif y_return_pct <= -FLAT_THR:
        direction = "DOWN"
    else:
        direction = "FLAT"

    # 일별 요약 (LLM 참고용)
    daily = []
    idx = y_df.index if hasattr(y_df.index, "strftime") else range(len(y_df))
    for i, (date, row) in enumerate(y_df.iterrows()):
        daily.append({
            "day":    i + 1,
            "date":   str(date.date()) if hasattr(date, "date") else str(date),
            "open":   float(row["open"]),
            "high":   float(row["high"]),
            "low":    float(row["low"]),
            "close":  float(row["close"]),
            "volume": int(row["volume"]) if "volume" in row else 0,
            "ret_pct": round(float((row["close"] / close0 - 1.0) * 100), 2),
        })

    return {
        "y_return_pct": round(y_return_pct, 3),
        "y_max_up":     round(y_max_up,     3),
        "y_max_down":   round(y_max_down,   3),
        "y_direction":  direction,
        "y_daily":      daily,
    }


def make_x_chart(x_df: pd.DataFrame) -> tuple[str, str]:
    """X 구간 텍스트 차트 + 메타데이터 반환"""
    indicators = {**add_ma(x_df, [20, 60]), **add_bollinger(x_df)}
    chart = plot_text_chart(
        x_df,
        rows=25,
        cols=X_DAYS,
        vol_rows=5,
        indicators=indicators,
        show_meta=False,
    )
    meta = generate_metadata(x_df)
    return chart, meta


def process_ticker(code: str, name: str) -> list[dict]:
    """단일 종목 → 슬라이딩 윈도우 페어 리스트"""
    pkl_path = CACHE_DIR / f"{code}.pkl"
    if not pkl_path.exists():
        log.warning(f"{code} {name}: 캐시 없음, 스킵")
        return []

    with open(pkl_path, "rb") as f:
        df = pickle.load(f)

    # 컬럼 정규화
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df.columns = df.columns.str.lower()
    # 중복 컬럼 제거 (yfinance가 복수 종목 데이터를 섞어서 반환하는 경우 대응)
    df = df.loc[:, ~df.columns.duplicated(keep="first")]
    df = df.dropna(subset=["open", "high", "low", "close"])
    df = df.sort_index()

    total = len(df)
    min_needed = X_DAYS + Y_DAYS

    if total < min_needed:
        log.warning(f"{code} {name}: 데이터 부족 ({total}일 < {min_needed}), 스킵")
        return []

    pairs = []
    # 슬라이딩 윈도우: X 시작 인덱스를 SHIFT씩 이동
    start_idx = 0
    while start_idx + min_needed <= total:
        x_end_idx = start_idx + X_DAYS
        y_end_idx = x_end_idx + Y_DAYS

        if y_end_idx > total:
            break

        x_df = df.iloc[start_idx:x_end_idx].copy()
        y_df = df.iloc[x_end_idx:y_end_idx].copy()

        # 날짜 문자열
        x_start_str = str(x_df.index[0].date())
        x_end_str   = str(x_df.index[-1].date())
        y_start_str = str(y_df.index[0].date())
        y_end_str   = str(y_df.index[-1].date())

        sample_id = f"{code}_{x_end_str}"

        try:
            chart, meta = make_x_chart(x_df)
            y_stats = compute_y_stats(y_df)
        except Exception as e:
            log.warning(f"{code} 윈도우 {x_end_str}: 오류 — {e}")
            start_idx += SHIFT
            continue

        pair = {
            "id":       sample_id,
            "ticker":   code,
            "name":     name,
            "x_start":  x_start_str,
            "x_end":    x_end_str,
            "y_start":  y_start_str,
            "y_end":    y_end_str,
            "x_chart":  chart,
            "x_meta":   meta,
            **y_stats,
        }
        pairs.append(pair)
        start_idx += SHIFT

    return pairs


def main():
    meta_map = load_ticker_meta()

    out_path = OUT_DIR / "pairs.jsonl"
    pkl_files = sorted(CACHE_DIR.glob("*.pkl"))

    if not pkl_files:
        log.error(f"캐시 없음. 먼저 01_download_ohlcv.py 실행하세요.")
        return

    log.info(f"캐시 종목: {len(pkl_files)}개 / 전체 목표: {len(meta_map)}개")

    total_pairs = 0
    skipped = 0

    with open(out_path, "w", encoding="utf-8") as f:
        for pkl_path in pkl_files:
            code = pkl_path.stem
            name = meta_map.get(code, code)

            pairs = process_ticker(code, name)
            if not pairs:
                skipped += 1
                continue

            for pair in pairs:
                f.write(json.dumps(pair, ensure_ascii=False) + "\n")
            total_pairs += len(pairs)

            log.info(f"✓ {code} {name}: {len(pairs)}개 페어")

    log.info(f"\n완료: 총 {total_pairs}개 페어 저장 → {out_path}")
    log.info(f"스킵: {skipped}종목")

    # 간단한 통계
    if total_pairs > 0:
        with open(out_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        directions = [json.loads(l)["y_direction"] for l in lines]
        up   = directions.count("UP")
        down = directions.count("DOWN")
        flat = directions.count("FLAT")
        log.info(f"방향 분포: UP={up} ({up/total_pairs*100:.1f}%) / "
                 f"DOWN={down} ({down/total_pairs*100:.1f}%) / "
                 f"FLAT={flat} ({flat/total_pairs*100:.1f}%)")


if __name__ == "__main__":
    main()
